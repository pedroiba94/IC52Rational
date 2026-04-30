# -*- coding: utf-8 -*-
"""
ic52ic_dialog.py — CAUMAX-style dialog for IC 5.2-IC Rational Method plugin.
QGIS 4.0 / Qt6 only.
"""

import os

from qgis.PyQt.QtCore import Qt, QThread, pyqtSignal
from qgis.PyQt.QtGui import QFont, QColor, QDoubleValidator, QIntValidator
from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QLineEdit, QComboBox, QPushButton, QTabWidget, QWidget,
    QGroupBox, QFrame, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QMessageBox, QSizePolicy, QSpacerItem,
    QCheckBox, QTextEdit, QFileDialog, QScrollArea,
)
from qgis.core import (
    QgsProject, QgsVectorLayer, QgsMapLayer, QgsFieldProxyModel,
    QgsMessageLog, Qgis,
    QgsCategorizedSymbolRenderer, QgsRendererCategory,
    QgsSymbol, QgsWkbTypes,
    QgsPalLayerSettings, QgsVectorLayerSimpleLabeling, QgsTextFormat,
)
from qgis.gui import QgsMapLayerComboBox, QgsFieldComboBox
from qgis.core import QgsMapLayerProxyModel

from .ic52ic_core import (
    CuencaInput, PrecipInput, SueloInput,
    calcular_metodo_racional, get_beta_data,
    OBRA_PM, OBRA_DT, NDIF, TABLE_2_5, LEVANTE_REGIONS,
    PERIODOS_STANDARD,
)

# ---------------------------------------------------------------------------
# ArcGIS Online layer URLs (BetaSpainMap service)
# Tab 2 (β regions) and Tab 3 (hydrological groups) use the same service
# until the hydro-groups layer is published separately.
# ---------------------------------------------------------------------------
ARCGIS_BETA_URL = (
    "https://services3.arcgis.com/XJ3WvEHgenh4hHTM"
    "/arcgis/rest/services/Regiones_Coeficientes_/FeatureServer/0"
)
ARCGIS_HIDRO_URL = (
    "https://services3.arcgis.com/XJ3WvEHgenh4hHTM"
    "/arcgis/rest/services/Grupos_Hidrologicos_IC52IC/FeatureServer/0"
)

# Region colours — same palette as BetaSpainMap (R, G, B, alpha 0-1)
_REGION_COLORS = {
    "11":   (215, 215, 215), "12":   (180, 215, 211), "13":   (255, 159, 162),
    "21":   (56,  170, 219), "22":   (252, 187, 161), "23":   (157,   0, 188),
    "24":   (249,  15, 120), "25":   (145, 172, 195), "31":   (255, 210, 221),
    "32":   (255, 246, 185), "33":   (233, 233, 233), "41":   (203, 178, 116),
    "42":   (220, 233, 174), "52":   (239, 187, 255), "53":   (0,   160, 201),
    "61":   (8,   101, 161), "71":   (137, 200,   0), "72":   (94,  202,  48),
    "81":   (255, 201, 227), "83":   (38,  102, 128), "91":   (76,  190,  84),
    "92":   (255, 173, 192), "93":   (153,  73, 137), "101":  (149,  73, 176),
    "511":  (53,  206,  36), "512":  (255, 202, 133), "821":  (150,  46, 198),
    "822":  (87,  216, 255), "941":  (220, 191, 113), "942":  (120, 191,  69),
    "951":  (219, 236, 255), "952":  (223, 192, 194), "1021": (175, 175, 168),
    "1022": (252, 254, 156),
}
_REGION_OPACITY = int(0.65 * 255)  # 65% — same as BetaSpainMap


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sep(parent=None):
    """Horizontal separator line."""
    sep = QFrame(parent)
    sep.setFrameShape((QFrame.Shape.HLine if hasattr(QFrame, "Shape") else QFrame.HLine))
    sep.setStyleSheet("color: #ccc;")
    return sep


def _label_bold(text, parent=None):
    lbl = QLabel(text, parent)
    f = lbl.font()
    f.setBold(True)
    lbl.setFont(f)
    return lbl


def _double_edit(placeholder="", parent=None):
    ed = QLineEdit(parent)
    ed.setPlaceholderText(placeholder)
    val = QDoubleValidator()
    val.setNotation(QDoubleValidator.Notation.StandardNotation)
    ed.setValidator(val)
    return ed


def _readonly_edit(parent=None):
    ed = QLineEdit(parent)
    ed.setReadOnly(True)
    ed.setStyleSheet(
        "QLineEdit { background-color: #f0f4f8; color: #333; border: 1px solid #ccc; }"
    )
    return ed


def _scrollable(widget):
    """Wrap a QWidget in a QScrollArea so tab content scrolls vertically."""
    scroll = QScrollArea()
    scroll.setWidgetResizable(True)
    scroll.setWidget(widget)
    scroll.setFrameShape((QFrame.Shape.NoFrame if hasattr(QFrame, "Shape") else QFrame.NoFrame))
    scroll.setHorizontalScrollBarPolicy((Qt.ScrollBarPolicy.ScrollBarAlwaysOff if hasattr(Qt, "ScrollBarPolicy") else Qt.ScrollBarAlwaysOff))
    scroll.setVerticalScrollBarPolicy((Qt.ScrollBarPolicy.ScrollBarAsNeeded if hasattr(Qt, "ScrollBarPolicy") else Qt.ScrollBarAsNeeded))
    return scroll


# ---------------------------------------------------------------------------
# P0i weighted calculation worker
# ---------------------------------------------------------------------------

class P0Worker(QThread):
    rows_ready = pyqtSignal(list, float)  # rows, total_area_km2
    error = pyqtSignal(str)

    def __init__(self, layer, grupo_hid, pendiente_umbral, parent=None):
        super().__init__(parent)
        self.layer = layer
        self.grupo_hid = grupo_hid
        self.pendiente_umbral = pendiente_umbral   # float %

    def run(self):
        """
        Returns a list of rows, one per unique CLC code found in the layer.
        Each row: (code_norm, code_raw, area_km2, P0i)
        Rows with unknown P0i are included with P0i=None for user awareness.
        """
        try:
            from .ic52ic_p0table import get_P0i_CLC, get_P0i_info, CLC_NAMES, _resolve_code
            # Aggregate area by normalized CLC code
            code_areas: dict = {}   # code_norm → total_area_km2
            code_raw: dict = {}     # code_norm → raw code string (for display)
            found_field = None

            for feat in self.layer.getFeatures():
                geom = feat.geometry()
                if geom is None or geom.isEmpty():
                    continue
                area = geom.area() / 1e6  # m² → km²

                fields = feat.fields().names()
                raw = ""
                for fname in ("CODE_18", "Code_18", "CLC_CODE", "code_18", "CODE18"):
                    if fname in fields:
                        raw = str(feat[fname])
                        found_field = fname
                        break

                # Normalize to 5-digit CLC code
                c = raw.strip()
                if c.isdigit():
                    if len(c) == 3:
                        c = c + "00"
                    elif len(c) == 4:
                        c = c + "0"

                if not c:
                    continue

                if c not in code_areas:
                    code_areas[c] = 0.0
                    code_raw[c] = raw
                code_areas[c] += area

            if not code_areas:
                all_fields = self.layer.fields().names()
                self.error.emit(
                    f"No se encontraron entidades válidas en la capa CLC.\n\n"
                    f"Campos disponibles: {all_fields}\n"
                    f"Se busca: CODE_18, Code_18 o CLC_CODE con valores tipo 211 o 21100."
                )
                return

            # Build result rows
            rows = []
            for code_norm, area in sorted(code_areas.items()):
                P0i, note = get_P0i_info(code_norm, self.pendiente_umbral, self.grupo_hid)
                rows.append((code_norm, code_raw[code_norm], area, P0i, note))

            total_area = sum(r[2] for r in rows)
            self.rows_ready.emit(rows, total_area)
        except Exception as e:
            self.error.emit(str(e))


# ---------------------------------------------------------------------------
# Main dialog
# ---------------------------------------------------------------------------

class IC52ICDialog(QDialog):
    def __init__(self, iface, parent=None):
        super().__init__(parent)
        self.iface = iface
        self.setWindowTitle("IC52Rational — Método Racional IC 5.2-IC")
        self.setMinimumWidth(700)
        self.setMinimumHeight(480)
        self.resize(860, 580)  # fits 1080p; tabs scroll internally
        # Slightly compact font for all child widgets
        app_font = self.font()
        app_font.setPointSize(max(8, app_font.pointSize() - 1))
        self.setFont(app_font)
        self.setWindowFlags(
            self.windowFlags()
            | (Qt.WindowType.Window if hasattr(Qt, 'WindowType') else Qt.Window)
            | (Qt.WindowType.WindowMaximizeButtonHint if hasattr(Qt, 'WindowType') else Qt.WindowMaximizeButtonHint)
            | (Qt.WindowType.WindowMinimizeButtonHint if hasattr(Qt, 'WindowType') else Qt.WindowMinimizeButtonHint)
            | (Qt.WindowType.WindowCloseButtonHint if hasattr(Qt, 'WindowType') else Qt.WindowCloseButtonHint)
        )
        self.setModal(False)  # non-modal: QGIS canvas stays interactive
        self._resultados = []
        self._layer_beta = None   # temporary ArcGIS layer for tab 2
        self._layer_hidro = None  # temporary ArcGIS layer for tab 3
        self._build_ui()
        # Connect tab change AFTER build so all widgets exist
        self._tabs.currentChanged.connect(self._on_tab_changed)

    # ------------------------------------------------------------------
    # Map layer management — show/hide ArcGIS layers per tab
    # ------------------------------------------------------------------

    def _on_tab_changed(self, index):
        """Show/hide contextual ArcGIS layers depending on active tab."""
        # Tab indices: 0=Cuenca, 1=Precipitación, 2=Suelo, 3=Resultados
        self._set_map_layer(1, index)  # β regions → tab 1 (Precipitación)
        self._set_map_layer(2, index)  # hidro groups → tab 2 (Suelo)

    def _set_map_layer(self, tab_index, current_index):
        """Load or toggle visibility of the ArcGIS layer for a given tab."""
        is_beta = (tab_index == 1)
        layer_attr = '_layer_beta' if is_beta else '_layer_hidro'
        url = ARCGIS_BETA_URL if is_beta else ARCGIS_HIDRO_URL
        layer_name = ('IC52IC — Regiones β (Fig. 2.9)'
                      if is_beta else 'IC52IC — Grupos Hidrológicos (Fig. 2.7)')

        layer = getattr(self, layer_attr)
        visible = (current_index == tab_index)
        if visible:
            # Load if not yet loaded or was removed from project
            if layer is None or QgsProject.instance().mapLayer(layer.id()) is None:
                layer = self._load_arcgis_layer(url, layer_name)
                setattr(self, layer_attr, layer)

            if layer:
                # Layer added with addMapLayer(False) — toggle via canvas layers
                canvas = self.iface.mapCanvas()
                cur_layers = list(canvas.layers())
                if layer not in cur_layers:
                    canvas.setLayers([layer] + cur_layers)
                canvas.refresh()
        else:
            if layer:
                canvas = self.iface.mapCanvas()
                canvas.setLayers([l for l in canvas.layers() if l.id() != layer.id()])
                canvas.refresh()

    def _load_arcgis_layer(self, feature_service_url, name):
        """
        Load an ArcGIS FeatureService layer using QGIS QgsNetworkAccessManager,
        which handles ArcGIS Online authentication tokens automatically.
        Falls back to urllib if the network manager is unavailable.
        """
        import urllib.parse, tempfile, os, json

        # API key selector — beta regions key vs hydro groups key
        _IS_HIDRO = "Hidro" in name or "hidro" in name or "Grupo" in name
        ARCGIS_API_KEY = (
            "AAPTaVMth54yHQ44Nr-02rX4oeg..QwW2MjeSCPNGQ4SEhcR21afVdjE_uKFQ7Mo_"
            "mhiTIG3S6XWGF7BkxN9sdJ_Xu2L4sTz6NS7804GOw8kViWZIsvIL2_j_wA6Hxe2-"
            "5cV_h2Gb4ibpOzsHqBipclDsh0jVFF8hHzcGwEIzGN7-ypYLzTLdKO74eimFenBVu"
            "iJYFR9Nc9hzqNf7FoANh7C0i-PMyedlDzAamFLl6TBwqgROFliEWRxTtQ..AT1_vPhrekCR"
        ) if _IS_HIDRO else (
            "AAPTaSYr1Oh9VcMKaXMiIw1LU0A..e_R7jPRsNKhCEZYmNa0JSPF3NU6Tf1FW"
            "vRcTAPWRpFOECZP8iUiOrvTAjj7UUgXt_sHc9LAi-S4OL46PHVTM2_kQOSmgEi"
            "GKnqanIU6ZZC2SXU-_RipVdlroyyeuVoRrElMNI9i7FV_279Q2T-JvvjVq9NRqy"
            "GnEScs_MlmFhWqcI-R2J6YsBYKuxWLDjgV6MiijnFR3Ba2dU45aLm-N1mXlqpnt"
            "Gw..AT1_0d2XVQ9Z"
        )

        params = urllib.parse.urlencode({
            "where": "1=1",
            "outFields": "*",
            "f": "geojson",
            "outSR": "4326",
            "resultRecordCount": "2000",
            "token": ARCGIS_API_KEY,
        })
        query_url = f"{feature_service_url}/query?{params}"
        data = None

        # ── Method 1: QGIS network manager (handles ArcGIS auth) ──────────
        try:
            from qgis.core import QgsNetworkAccessManager
            from qgis.PyQt.QtNetwork import QNetworkRequest
            from qgis.PyQt.QtCore import QUrl, QEventLoop
            from qgis.PyQt.QtNetwork import QNetworkReply

            manager = QgsNetworkAccessManager.instance()
            request = QNetworkRequest(QUrl(query_url))
            request.setRawHeader(b"User-Agent", b"QGIS-IC52IC/1.0")

            loop = QEventLoop()
            reply = manager.get(request)
            reply.finished.connect(loop.quit)
            loop.exec()

            if reply.error() == (QNetworkReply.NetworkError.NoError if hasattr(QNetworkReply, "NetworkError") else QNetworkReply.NoError):
                data = bytes(reply.readAll()).decode("utf-8")
            else:
                QgsMessageLog.logMessage(
                    f"IC52IC: Network error ({reply.error()}): {reply.errorString()}",
                    "IC52IC", Qgis.MessageLevel.Warning
                )
            reply.deleteLater()
        except Exception as e:
            QgsMessageLog.logMessage(
                f"IC52IC: QgsNetworkAccessManager failed ({e}), trying urllib",
                "IC52IC", Qgis.MessageLevel.Warning
            )

        # ── Method 2: urllib fallback ──────────────────────────────────────
        if data is None:
            try:
                import urllib.request
                req = urllib.request.Request(
                    query_url, headers={"User-Agent": "QGIS-IC52IC/1.0"}
                )
                with urllib.request.urlopen(req, timeout=30) as resp:
                    data = resp.read().decode("utf-8")
            except Exception as e2:
                QgsMessageLog.logMessage(
                    f"IC52IC: urllib fallback failed: {e2}",
                    "IC52IC", Qgis.MessageLevel.Warning
                )

        if data is None:
            self.iface.messageBar().pushMessage(
                "IC52IC",
                f"No se pudo cargar '{name}'. Comprueba la conexión a internet.",
                level=Qgis.MessageLevel.Warning, duration=6
            )
            return None

        # ── Validate and save GeoJSON ─────────────────────────────────────
        try:
            parsed = json.loads(data)
            if "error" in parsed:
                raise ValueError(f"ArcGIS error: {parsed['error']}")
            if "features" not in parsed:
                raise ValueError("Respuesta sin 'features'")
        except (json.JSONDecodeError, ValueError) as e:
            QgsMessageLog.logMessage(
                f"IC52IC: GeoJSON inválido para '{name}': {e} | primeros 300: {data[:300]}",
                "IC52IC", Qgis.MessageLevel.Warning
            )
            self.iface.messageBar().pushMessage(
                "IC52IC",
                f"Respuesta inválida del servidor ArcGIS para '{name}'.",
                level=Qgis.MessageLevel.Warning, duration=6
            )
            return None

        tmp = tempfile.NamedTemporaryFile(
            suffix=".geojson", delete=False, mode="w", encoding="utf-8"
        )
        json.dump(parsed, tmp, ensure_ascii=False)
        tmp_path = tmp.name
        tmp.close()

        layer = QgsVectorLayer(tmp_path, name, "ogr")
        if not layer.isValid():
            QgsMessageLog.logMessage(
                f"IC52IC: Capa no válida: {name}",
                "IC52IC", Qgis.MessageLevel.Warning
            )
            return None

        # Apply style depending on layer type
        if 'Hidro' in name or 'hidro' in name or 'GRUPO' in [f.name() for f in layer.fields()]:
            self._apply_hidro_style(layer)
        else:
            self._apply_region_style(layer)
            self._apply_region_labels(layer)

        # addMapLayer(False) → renders on canvas, NOT shown in Layers panel
        QgsProject.instance().addMapLayer(layer, False)
        QgsMessageLog.logMessage(
            f"IC52IC: Capa '{name}' cargada — {layer.featureCount()} entidades",
            "IC52IC", Qgis.MessageLevel.Info
        )
        return layer

    def _apply_hidro_style(self, layer):
        """Categorized style for hydrological groups layer (B=golden, C=cream)."""
        try:
            field_names = [f.name() for f in layer.fields()]
            grupo_field = next(
                (f for f in ("GRUPO", "grupo", "HYDRO_GRP", "GROUP") if f in field_names),
                field_names[0] if field_names else "GRUPO"
            )
            HIDRO_COLORS = {
                "B": QColor(245, 235, 190, 180),   # crema claro — Grupo B (zona predominante)
                "C": QColor(210, 170,  80, 180),   # dorado/ocre — Grupo C (manchas)
            }
            categories = []
            for grupo, color in HIDRO_COLORS.items():
                symbol = QgsSymbol.defaultSymbol((QgsWkbTypes.GeometryType.PolygonGeometry if hasattr(QgsWkbTypes, "GeometryType") else QgsWkbTypes.PolygonGeometry))
                symbol.setColor(color)
                symbol.symbolLayer(0).setStrokeColor(QColor(120, 80, 20, 180))
                symbol.symbolLayer(0).setStrokeWidth(0.4)
                cat = QgsRendererCategory(grupo, symbol, f"Grupo Hidrológico {grupo}")
                categories.append(cat)
            renderer = QgsCategorizedSymbolRenderer(grupo_field, categories)
            layer.setRenderer(renderer)
            # Labels
            from qgis.core import QgsTextBackgroundSettings
            from qgis.PyQt.QtCore import QSizeF
            settings = QgsPalLayerSettings()
            settings.fieldName = grupo_field
            settings.enabled = True
            text_format = QgsTextFormat()
            text_format.setSize(8)
            # White text with dark stroke for visibility on both colours
            text_format.setColor(QColor(255, 255, 255, 255))
            settings.setFormat(text_format)
            layer.setLabelsEnabled(True)
            layer.setLabeling(QgsVectorLayerSimpleLabeling(settings))
            layer.triggerRepaint()
        except Exception as e:
            QgsMessageLog.logMessage(
                f"IC52IC: error estilo hidro: {e}", "IC52IC", Qgis.MessageLevel.Warning
            )

    def _apply_region_style(self, layer):
        """Apply BetaSpainMap-identical categorized colour style."""
        try:
            field_names = [f.name() for f in layer.fields()]
            region_field = next(
                (f for f in ("REGION", "region", "COD_REGION", "CODIGO")
                 if f in field_names),
                field_names[0] if field_names else "REGION"
            )
            categories = []
            for region_str, (r, g, b) in _REGION_COLORS.items():
                symbol = QgsSymbol.defaultSymbol((QgsWkbTypes.GeometryType.PolygonGeometry if hasattr(QgsWkbTypes, "GeometryType") else QgsWkbTypes.PolygonGeometry))
                color = QColor(r, g, b, _REGION_OPACITY)
                symbol.setColor(color)
                symbol.symbolLayer(0).setStrokeColor(QColor(35, 35, 35, 255))
                symbol.symbolLayer(0).setStrokeWidth(0.3)
                cat = QgsRendererCategory(region_str, symbol, region_str)
                categories.append(cat)
            renderer = QgsCategorizedSymbolRenderer(region_field, categories)
            layer.setRenderer(renderer)
            layer.triggerRepaint()
        except Exception as e:
            QgsMessageLog.logMessage(
                f"IC52IC: error estilo: {e}", "IC52IC", Qgis.MessageLevel.Warning
            )

    def _apply_region_labels(self, layer):
        """Apply region code labels with white background box."""
        try:
            from qgis.core import QgsTextBackgroundSettings
            from qgis.PyQt.QtCore import QSizeF
            field_names = [f.name() for f in layer.fields()]
            region_field = next(
                (f for f in ("REGION", "region", "COD_REGION", "CODIGO")
                 if f in field_names),
                field_names[0] if field_names else "REGION"
            )
            settings = QgsPalLayerSettings()
            settings.fieldName = region_field
            settings.enabled = True
            text_format = QgsTextFormat()
            text_format.setSize(8)
            text_format.setColor(QColor(0, 0, 0, 255))
            bg = QgsTextBackgroundSettings()
            bg.setEnabled(True)
            bg.setType(
                (QgsTextBackgroundSettings.ShapeType.ShapeRectangle if hasattr(QgsTextBackgroundSettings, "ShapeType") else QgsTextBackgroundSettings.ShapeRectangle)
                if hasattr(QgsTextBackgroundSettings, 'ShapeType')
                else QgsTextBackgroundSettings.ShapeRectangle
            )
            bg.setFillColor(QColor(255, 255, 255, 220))
            bg.setStrokeColor(QColor(0, 0, 0, 255))
            bg.setStrokeWidth(0.3)
            bg.setSizeType(
                (QgsTextBackgroundSettings.SizeType.SizeBuffer if hasattr(QgsTextBackgroundSettings, "SizeType") else QgsTextBackgroundSettings.SizeBuffer)
                if hasattr(QgsTextBackgroundSettings, 'SizeType')
                else QgsTextBackgroundSettings.SizeBuffer
            )
            bg.setSize(QSizeF(1.0, 0.5))
            text_format.setBackground(bg)
            settings.setFormat(text_format)
            layer.setLabelsEnabled(True)
            layer.setLabeling(QgsVectorLayerSimpleLabeling(settings))
        except Exception as e:
            QgsMessageLog.logMessage(
                f"IC52IC: error etiquetas: {e}", "IC52IC", Qgis.MessageLevel.Warning
            )

    def _cleanup_map_layers(self):
        """Remove temporary IC52IC layers when dialog closes."""
        for attr in ('_layer_beta', '_layer_hidro'):
            layer = getattr(self, attr)
            if layer and QgsProject.instance().mapLayer(layer.id()):
                QgsProject.instance().removeMapLayer(layer.id())
            setattr(self, attr, None)

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(6)

        # ── Header ──────────────────────────────────────────────────────
        hdr = QLabel(
            "<b style='font-size:13px;'>Cálculo de caudales — Método Racional IC 5.2-IC</b>"
            "<span style='color:#666; font-size:10px;'>  |  Capítulo 2</span>"
        )
        hdr.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        hdr.setStyleSheet(
            "background-color: #1a5276; color: white; padding: 8px 12px; border-radius: 4px;"
        )
        main_layout.addWidget(hdr)

        # ── Nombre del proyecto ──────────────────────────────────────────
        hbox_nombre = QHBoxLayout()
        hbox_nombre.addWidget(QLabel("Nombre del proyecto / cuenca:"))
        self._ed_nombre = QLineEdit()
        self._ed_nombre.setPlaceholderText("Ejemplo: Cuenca del arroyo de la Degollada")
        hbox_nombre.addWidget(self._ed_nombre)
        main_layout.addLayout(hbox_nombre)

        # ── Tabs ─────────────────────────────────────────────────────────
        self._tabs = QTabWidget()
        self._tabs.addTab(self._tab_cuenca(), "① Cuenca")
        self._tabs.addTab(self._tab_precipitacion(), "② Precipitación")
        self._tabs.addTab(self._tab_suelo(), "③ Suelo / P₀")
        self._tabs.addTab(self._tab_resultados(), "④ Resultados")
        main_layout.addWidget(self._tabs)

        # ── Bottom buttons ────────────────────────────────────────────────
        main_layout.addWidget(_sep())
        btn_layout = QHBoxLayout()
        self._btn_calcular = QPushButton("▶  Calcular")
        self._btn_calcular.setMinimumHeight(36)
        self._btn_calcular.setStyleSheet(
            "QPushButton { background-color: #1a5276; color: white; font-weight: bold; "
            "border-radius: 4px; padding: 4px 20px; }"
            "QPushButton:hover { background-color: #2471a3; }"
        )
        self._btn_calcular.clicked.connect(self._on_calcular)

        self._btn_limpiar = QPushButton("Limpiar")
        self._btn_limpiar.clicked.connect(self._on_limpiar)

        self._btn_cerrar = QPushButton("Cerrar")
        self._btn_cerrar.clicked.connect(self._on_cerrar)

        btn_layout.addWidget(self._btn_calcular)
        btn_layout.addStretch()
        btn_layout.addWidget(self._btn_limpiar)
        btn_layout.addWidget(self._btn_cerrar)
        main_layout.addLayout(btn_layout)

    # ------------------------------------------------------------------
    # Tab 1 — Cuenca
    # ------------------------------------------------------------------

    def _tab_cuenca(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setSpacing(10)

        # Geometría
        grp_geom = QGroupBox("Geometría de la cuenca")
        form = QFormLayout(grp_geom)
        form.setRowWrapPolicy(QFormLayout.RowWrapPolicy.WrapAllRows)

        self._ed_area = _double_edit("km²")
        self._ed_lc   = _double_edit("km")
        self._ed_jc   = _double_edit("m/m — ejemplo: 0.015")
        form.addRow("Área A (km²):", self._ed_area)
        form.addRow("Longitud cauce Lc (km):", self._ed_lc)
        form.addRow("Pendiente media cauce Jc (m/m):", self._ed_jc)

        # Variables calculadas (readonly)
        self._ed_KA  = _readonly_edit()
        self._ed_tc  = _readonly_edit()
        self._ed_Kt  = _readonly_edit()
        form.addRow("Factor reductor KA:", self._ed_KA)
        form.addRow("Tiempo de concentración tc (h):", self._ed_tc)
        form.addRow("Coef. uniformidad temporal Kt:", self._ed_Kt)

        btn_calc_tc = QPushButton("Calcular KA / tc / Kt")
        btn_calc_tc.clicked.connect(self._on_calc_tc)
        form.addRow("", btn_calc_tc)
        layout.addWidget(grp_geom)

        # Tipo de cuenca
        grp_tipo = QGroupBox("Tipo de cuenca")
        hbox = QHBoxLayout(grp_tipo)
        self._cb_tipo_cuenca = QComboBox()
        self._cb_tipo_cuenca.addItems(["Principal (fórmula §2.2.2.5)", "Secundaria (flujo difuso §2.2.2.5)"])
        self._cb_tipo_cuenca.currentIndexChanged.connect(self._on_tipo_cuenca_changed)
        hbox.addWidget(self._cb_tipo_cuenca)
        layout.addWidget(grp_tipo)

        # Flujo difuso (oculto por defecto)
        self._grp_difuso = QGroupBox("Parámetros flujo difuso (cuenca secundaria)")
        self._grp_difuso.setVisible(False)
        form_dif = QFormLayout(self._grp_difuso)
        self._ed_ldif  = _double_edit("m")
        self._cb_ndif  = QComboBox()
        self._cb_ndif.addItems([
            "Pavimentado o revestido (0.015)",
            "Sin vegetación (0.050)",
            "Vegetación escasa (0.120)",
            "Vegetación media (0.320)",
            "Vegetación densa (1.000)",
        ])
        self._ed_jdif = _double_edit("m/m")
        form_dif.addRow("Longitud difuso Ldif (m):", self._ed_ldif)
        form_dif.addRow("Cobertura terreno (ndif):", self._cb_ndif)
        form_dif.addRow("Pendiente difuso Jdif (m/m):", self._ed_jdif)
        layout.addWidget(self._grp_difuso)

        layout.addStretch()
        return _scrollable(w)

    # ------------------------------------------------------------------
    # Tab 2 — Precipitación
    # ------------------------------------------------------------------

    def _tab_precipitacion(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setSpacing(10)

        # ── Cv / P calculator ──────────────────────────────────────────────
        grp_cv = QGroupBox("Calcular Pd desde P media y Cv — Publicación DGC (opcional)")
        grp_cv.setCheckable(True)
        grp_cv.setChecked(False)
        self._grp_cv = grp_cv
        form_cv = QFormLayout(grp_cv)

        self._ed_P_media = _double_edit("mm — precipitación media diaria máxima anual")
        self._ed_Cv      = _double_edit("adim — coef. variación (0.30 – 0.52)")
        form_cv.addRow("P media (mm):", self._ed_P_media)
        form_cv.addRow("Cv (adim.):", self._ed_Cv)

        btn_calc_pd = QPushButton("▶  Calcular Pd para todos los T")
        btn_calc_pd.setStyleSheet(
            "QPushButton { background:#1a5276; color:white; font-weight:bold; "
            "border-radius:4px; padding:4px 12px; }"
            "QPushButton:hover { background:#2471a3; }"
        )
        btn_calc_pd.clicked.connect(self._on_calc_pd_cv)
        form_cv.addRow("", btn_calc_pd)

        lbl_cv = QLabel(
            "<i style='color:#555; font-size:10px;'>"
            "Pd(T) = P_media × YT(Cv, T) — interpolación lineal en tabla DGC.<br>"
            "Los valores se copian automáticamente a los campos Pd de arriba.<br>"
            "Cv válido: 0.30 – 0.52. T disponibles: 2, 5, 10, 25, 50, 100, 500.<br>"
            "MCO (T≈2.5): interpolado entre T=2 y T=5.</i>"
        )
        lbl_cv.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        lbl_cv.setWordWrap(True)
        form_cv.addRow("", lbl_cv)
        layout.addWidget(grp_cv)

        # Precipitación diaria por T
        grp_pd = QGroupBox("Precipitación diaria Pd (mm) — por período de retorno")
        form_pd = QFormLayout(grp_pd)

        # 8 periods: 2, MCO(2.5), 5, 10, 25, 50, 100, 500
        self._pd_editors = {}
        pd_labels = [
            (2,     "T = 2 años"),
            ("MCO", "MCO (≈ 2.5 años)"),
            (5,     "T = 5 años"),
            (10,    "T = 10 años"),
            (25,    "T = 25 años"),
            (50,    "T = 50 años"),
            (100,   "T = 100 años"),
            (500,   "T = 500 años"),
        ]
        for T_key, label in pd_labels:
            ed = _double_edit("mm")
            form_pd.addRow(f"Pd {label} (mm):", ed)
            self._pd_editors[T_key] = ed

        lbl_pd = QLabel(
            "<i style='color:#666; font-size:10px;'>"
            "Fuente: Mapas DGC (publicación MOPU) o estudio estadístico pluviométrico. "
            "Adoptar el mayor valor obtenido. "
            "MCO = Máxima Crecida Ordinaria, T ≈ 2.5 años (interpolado).</i>"
        )
        lbl_pd.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        lbl_pd.setWordWrap(True)
        form_pd.addRow("", lbl_pd)
        layout.addWidget(grp_pd)

        # Índice de torrencialidad
        grp_tor = QGroupBox("Índice de torrencialidad I₁/Id — Figura 2.4")
        form_tor = QFormLayout(grp_tor)
        self._ed_I1Id = _double_edit("adim — típico: 8, 9, 10, 11, 12")
        form_tor.addRow("I₁/Id:", self._ed_I1Id)
        lbl_tor = QLabel(
            "<i style='color:#666; font-size:10px;'>"
            "Valor obtenido del mapa de la Figura 2.4 (IC 5.2-IC).<br>"
            "Burgos/Castilla y León: zona I₁/Id = 9 − 10.<br>"
            "Para más información consulta el mapa de isolíneas I₁/Id: "
            '<a href="https://www.transportes.gob.es/recursos_mfom/0610300.pdf">'
            "Publicación DGC (p. 19/55)</a>.</i>"
        )
        lbl_tor.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        lbl_tor.setOpenExternalLinks(True)
        lbl_tor.setWordWrap(True)
        form_tor.addRow("", lbl_tor)
        layout.addWidget(grp_tor)

        # Región β
        grp_beta = QGroupBox("Región — Coeficiente corrector β (Figura 2.9)")
        form_beta = QFormLayout(grp_beta)
        self._cb_region = QComboBox()
        regiones = sorted(TABLE_2_5.keys())
        for r in regiones:
            d = TABLE_2_5[r]
            label = f"Región {r}  (βm={d['bm']:.2f})"
            if r in LEVANTE_REGIONS:
                label += "  ⚠ Levante/SE"
            self._cb_region.addItem(label, r)
        # Default to region 32 (central Spain)
        idx32 = [self._cb_region.itemData(i) for i in range(self._cb_region.count())].index(32)
        self._cb_region.setCurrentIndex(idx32)
        form_beta.addRow("Región (Fig. 2.9):", self._cb_region)

        self._cb_tipo_obra = QComboBox()
        self._cb_tipo_obra.addItem("Drenaje transversal carretera — βDT = (βm − Δ50)·FT", OBRA_DT)
        self._cb_tipo_obra.addItem("Plataforma/márgenes o vías auxiliares — βPM = βm·FT", OBRA_PM)
        form_beta.addRow("Tipo de obra:", self._cb_tipo_obra)

        # Preview beta values
        self._tbl_beta = QTableWidget(1, 4)
        self._tbl_beta.setHorizontalHeaderLabels(["βm", "Δ50", "T=10 (ref)", "Tipo"])
        self._tbl_beta.setMaximumHeight(58)
        self._tbl_beta.horizontalHeader().setSectionResizeMode(
            (QHeaderView.ResizeMode.Stretch if hasattr(QHeaderView, "ResizeMode") else QHeaderView.Stretch)
        )
        self._tbl_beta.setEditTriggers((QAbstractItemView.EditTrigger.NoEditTriggers if hasattr(QAbstractItemView, "EditTrigger") else QAbstractItemView.NoEditTriggers))
        self._tbl_beta.verticalHeader().setVisible(False)
        form_beta.addRow("", self._tbl_beta)
        self._cb_region.currentIndexChanged.connect(self._on_region_changed)
        self._cb_tipo_obra.currentIndexChanged.connect(self._on_region_changed)
        self._on_region_changed()
        layout.addWidget(grp_beta)

        # IDF opcional
        grp_idf = QGroupBox("Curvas IDF (opcional — Figura 2.5)")
        grp_idf.setCheckable(True)
        grp_idf.setChecked(False)
        form_idf = QFormLayout(grp_idf)
        self._grp_idf = grp_idf
        self._ed_IIDF_tc  = _double_edit("mm/h")
        self._ed_IIDF_24  = _double_edit("mm/h")
        self._ed_kb       = _double_edit("adim — defecto 1.13")
        self._ed_kb.setText("1.13")
        form_idf.addRow("I_IDF(T, tc) (mm/h):", self._ed_IIDF_tc)
        form_idf.addRow("I_IDF(T, 24) (mm/h):", self._ed_IIDF_24)
        form_idf.addRow("kb:", self._ed_kb)
        layout.addWidget(grp_idf)

        layout.addStretch()
        return _scrollable(w)

    # ------------------------------------------------------------------
    # Tab 3 — Suelo / P0i
    # ------------------------------------------------------------------

    def _tab_suelo(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setSpacing(8)

        # ── Grupo hidrológico ────────────────────────────────────────────────
        grp_hid = QGroupBox("Grupo hidrológico de suelo — Tabla 2.4 / Figura 2.7")
        form_hid = QFormLayout(grp_hid)
        self._cb_grupo_hid = QComboBox()
        self._cb_grupo_hid.addItems([
            "A — Infiltración rápida (arenoso, perfecto drenaje)",
            "B — Infiltración moderada (franco-arenoso, buen drenaje)",
            "C — Infiltración lenta (franco-arcilloso, drenaje imperfecto)",
            "D — Infiltración muy lenta (arcilloso, drenaje pobre)",
        ])
        self._cb_grupo_hid.currentIndexChanged.connect(self._on_grupo_hid_changed)
        form_hid.addRow("Grupo hidrológico:", self._cb_grupo_hid)
        lbl_hid_pdf = QLabel(
            '<i style="color:#555; font-size:10px;">'
            'Para consultar el mapa de grupos hidrológicos (Fig. 2.7) ver '
            '<a href="https://www.boe.es/boe/dias/2016/03/10/pdfs/BOE-A-2016-2405.pdf">'
            'IC 5.2-IC (BOE 2016) p. 25/142</a>.</i>'
        )
        lbl_hid_pdf.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        lbl_hid_pdf.setOpenExternalLinks(True)
        lbl_hid_pdf.setWordWrap(True)
        form_hid.addRow("", lbl_hid_pdf)
        layout.addWidget(grp_hid)

        # ── Capa CLC ─────────────────────────────────────────────────────────
        grp_clc = QGroupBox("Capa CLC cargada en QGIS (Tabla 2.3)")
        form_clc = QFormLayout(grp_clc)

        self._cb_capa_clc = QgsMapLayerComboBox()
        self._cb_capa_clc.setFilters(
            QgsMapLayerProxyModel.Filter.VectorLayer
            if hasattr(QgsMapLayerProxyModel, "Filter")
            else QgsMapLayerProxyModel.VectorLayer
        )
        form_clc.addRow("Capa CLC:", self._cb_capa_clc)

        self._ed_pendiente_umbral = _double_edit("% — pendiente media cuenca para separar <3% / ≥3%")
        self._ed_pendiente_umbral.setText("3.0")
        form_clc.addRow("Pendiente media cuenca (%):", self._ed_pendiente_umbral)

        btn_calc_p0 = QPushButton("▶  Cargar usos de suelo desde CLC")
        btn_calc_p0.setStyleSheet(
            "QPushButton { background:#1a5276; color:white; font-weight:bold; "
            "border-radius:4px; padding:4px 12px; }"
            "QPushButton:hover { background:#2471a3; }"
        )
        btn_calc_p0.clicked.connect(self._on_calc_p0_clc)
        form_clc.addRow("", btn_calc_p0)
        layout.addWidget(grp_clc)

        # ── Tabla desglosada por uso de suelo ────────────────────────────────
        grp_tbl = QGroupBox("Desglose P0i por uso de suelo (Tabla 2.3) — editable")
        tbl_layout = QVBoxLayout(grp_tbl)

        lbl_info = QLabel(
            "<i style='color:#555; font-size:10px;'>"
            "El cálculo aplica §2.2.4 (cuenca heterogénea): QT = Kt/3.6 · I(T,tc) · Σ[Ci·Ai]<br>"
            "Puedes editar P0i manualmente si alguna fila necesita ajuste.</i>"
        )
        lbl_info.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        lbl_info.setWordWrap(True)
        tbl_layout.addWidget(lbl_info)

        self._tbl_clc = QTableWidget(0, 6)
        self._tbl_clc.setHorizontalHeaderLabels([
            "Código CLC", "Uso de suelo", "Grupo Hid.", "Pendiente", "Área (km²)", "P0i (mm)",
        ])
        self._tbl_clc.horizontalHeader().setSectionResizeMode((QHeaderView.ResizeMode.Stretch if hasattr(QHeaderView, "ResizeMode") else QHeaderView.Stretch))
        self._tbl_clc.horizontalHeader().setSectionResizeMode(0, (QHeaderView.ResizeMode.ResizeToContents if hasattr(QHeaderView, "ResizeMode") else QHeaderView.ResizeToContents))
        self._tbl_clc.horizontalHeader().setSectionResizeMode(2, (QHeaderView.ResizeMode.ResizeToContents if hasattr(QHeaderView, "ResizeMode") else QHeaderView.ResizeToContents))
        self._tbl_clc.horizontalHeader().setSectionResizeMode(3, (QHeaderView.ResizeMode.ResizeToContents if hasattr(QHeaderView, "ResizeMode") else QHeaderView.ResizeToContents))
        self._tbl_clc.setMinimumHeight(180)
        self._tbl_clc.setSelectionMode((QAbstractItemView.SelectionMode.SingleSelection if hasattr(QAbstractItemView, "SelectionMode") else QAbstractItemView.SingleSelection))
        # P0i column (last) is editable; others are readonly
        tbl_layout.addWidget(self._tbl_clc)

        # Summary row
        hbox_sum = QHBoxLayout()
        hbox_sum.addWidget(QLabel("Área total cargada:"))
        self._ed_area_total_clc = _readonly_edit()
        self._ed_area_total_clc.setMaximumWidth(110)
        hbox_sum.addWidget(self._ed_area_total_clc)
        hbox_sum.addWidget(QLabel("km²"))
        hbox_sum.addStretch()
        lbl_note = QLabel(
            "<i style='color:#888; font-size:10px;'>"
            "β se aplica automáticamente en el cálculo. P0 final = P0i · β(región, T, tipo_obra)</i>"
        )
        lbl_note.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        hbox_sum.addWidget(lbl_note)
        tbl_layout.addLayout(hbox_sum)
        layout.addWidget(grp_tbl)

        # ── P0i manual override ───────────────────────────────────────────────
        grp_manual = QGroupBox("P0i manual (opcional — si no usas capa CLC)")
        grp_manual.setCheckable(True)
        grp_manual.setChecked(False)
        self._grp_p0_manual = grp_manual
        form_m = QFormLayout(grp_manual)
        self._ed_P0i_manual = _double_edit("mm")
        form_m.addRow("P0i única para toda la cuenca (mm):", self._ed_P0i_manual)
        lbl_m = QLabel(
            "<i style='color:#888; font-size:10px;'>"
            "Si está activo, se ignora la tabla CLC y se usa un P0i único (cuenca homogénea).</i>"
        )
        lbl_m.setTextFormat((Qt.TextFormat.RichText if hasattr(Qt, "TextFormat") else Qt.RichText))
        form_m.addRow("", lbl_m)
        layout.addWidget(grp_manual)

        return _scrollable(w)

    # ------------------------------------------------------------------
    # Tab 4 — Resultados
    # ------------------------------------------------------------------

    def _tab_resultados(self):
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setSpacing(8)

        # Variables intermedias (common to all T)
        grp_inter = QGroupBox("Variables intermedias (comunes a todos los T)")
        form_i = QGridLayout(grp_inter)

        lbl_names = ["KA", "tc (h)", "Kt", "I₁/Id", "P0i (mm)"]
        self._res_eds = {}
        for col, name in enumerate(lbl_names):
            form_i.addWidget(QLabel(name + ":"), 0, col * 2)
            ed = _readonly_edit()
            ed.setMaximumWidth(90)
            form_i.addWidget(ed, 0, col * 2 + 1)
            self._res_eds[name] = ed
        layout.addWidget(grp_inter)

        # Results table T=10, 100, 500
        grp_tbl = QGroupBox("Resultados por período de retorno")
        tbl_layout = QVBoxLayout(grp_tbl)

        self._tbl_res = QTableWidget(0, 10)
        self._tbl_res.setHorizontalHeaderLabels([
            "T (años)", "Pd (mm)", "Id (mm/h)", "Fint", "I (mm/h)",
            "β", "P0 (mm)", "C", "Σ(Ci·Ai)", "QT (m³/s)",
        ])
        self._tbl_res.horizontalHeader().setSectionResizeMode(
            (QHeaderView.ResizeMode.Stretch if hasattr(QHeaderView, "ResizeMode") else QHeaderView.Stretch)
        )
        self._tbl_res.setEditTriggers((QAbstractItemView.EditTrigger.NoEditTriggers if hasattr(QAbstractItemView, "EditTrigger") else QAbstractItemView.NoEditTriggers))
        self._tbl_res.setSelectionMode((QAbstractItemView.SelectionMode.SingleSelection if hasattr(QAbstractItemView, "SelectionMode") else QAbstractItemView.SingleSelection))
        self._tbl_res.setMinimumHeight(130)
        tbl_layout.addWidget(self._tbl_res)
        layout.addWidget(grp_tbl)

        # QT highlight boxes — 8 periods in 2 rows of 4
        grp_q = QGroupBox("Caudales de diseño QT (m³/s)")
        vbox_q = QVBoxLayout(grp_q)
        self._q_boxes = {}

        row1_T = [
            (2,     "#85c1e9", "T=2"),
            ("MCO", "#5dade2", "MCO"),
            (5,     "#2e86c1", "T=5"),
            (10,    "#2ecc71", "T=10"),
        ]
        row2_T = [
            (25,  "#f39c12", "T=25"),
            (50,  "#e67e22", "T=50"),
            (100, "#cb4335", "T=100"),
            (500, "#922b21", "T=500"),
        ]

        for row_data in [row1_T, row2_T]:
            hbox_q = QHBoxLayout()
            for T_key, color, label in row_data:
                vb = QVBoxLayout()
                lbl_t = QLabel(label)
                lbl_t.setAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
                lbl_t.setStyleSheet(
                    f"background:{color}; color:white; font-weight:bold; "
                    f"border-radius:4px; padding:3px;"
                )
                ed_q = QLineEdit()
                ed_q.setReadOnly(True)
                ed_q.setAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
                ed_q.setStyleSheet(
                    f"font-size:13px; font-weight:bold; border:2px solid {color}; "
                    f"border-radius:4px; padding:4px;"
                )
                vb.addWidget(lbl_t)
                vb.addWidget(ed_q)
                hbox_q.addLayout(vb)
                self._q_boxes[T_key] = ed_q
            vbox_q.addLayout(hbox_q)
        layout.addWidget(grp_q)

        # Warnings
        self._txt_warnings = QTextEdit()
        self._txt_warnings.setReadOnly(True)
        self._txt_warnings.setMaximumHeight(80)
        self._txt_warnings.setPlaceholderText("Avisos y notas aparecerán aquí...")
        layout.addWidget(QLabel("Avisos:"))
        layout.addWidget(self._txt_warnings)

        # ── Export button ────────────────────────────────────────────────
        btn_export = QPushButton("📥  Exportar resultados a Excel (.xlsx)")
        btn_export.setMinimumHeight(34)
        btn_export.setStyleSheet(
            "QPushButton { background:#1e8449; color:white; font-weight:bold; "
            "border-radius:4px; padding:4px 16px; }"
            "QPushButton:hover { background:#27ae60; }"
        )
        btn_export.clicked.connect(self._on_export_excel)
        layout.addWidget(btn_export)

        return _scrollable(w)

    # ------------------------------------------------------------------
    # Signal handlers
    # ------------------------------------------------------------------

    def _on_tipo_cuenca_changed(self, idx):
        self._grp_difuso.setVisible(idx == 1)

    def _on_grupo_hid_changed(self):
        """Re-read P0i for all rows when hydrological group changes."""
        if self._tbl_clc.rowCount() == 0:
            return
        from .ic52ic_p0table import get_P0i_CLC, get_P0i_info
        grupo_map = {0: "A", 1: "B", 2: "C", 3: "D"}
        grupo = grupo_map[self._cb_grupo_hid.currentIndex()]
        try:
            pend = float(self._ed_pendiente_umbral.text().replace(",", "."))
        except ValueError:
            pend = 3.0
        pend_label = f"{'≥' if pend >= 3.0 else '<'}3%"

        colors_ok  = QColor("#d5f5e3")
        colors_err = QColor("#fadbd8")
        for row in range(self._tbl_clc.rowCount()):
            code = self._tbl_clc.item(row, 0).text()
            P0i, note = get_P0i_info(code, pend, grupo)
            bg = colors_ok if P0i is not None else colors_err
            for col in range(self._tbl_clc.columnCount()):
                if self._tbl_clc.item(row, col):
                    self._tbl_clc.item(row, col).setBackground(bg)
            # Update grupo col
            self._tbl_clc.item(row, 2).setText(grupo)
            self._tbl_clc.item(row, 3).setText(pend_label)
            # Update P0i col
            p0i_val = f"{P0i:.1f}" if P0i is not None else "N/D"
            it = self._tbl_clc.item(row, 5)
            it.setText(p0i_val)
            if note:
                it.setToolTip(f"⚠ {note}")
                it.setForeground(QColor("#7d6608"))
            else:
                it.setToolTip("")
                it.setForeground(QColor("#000000"))

    def _on_region_changed(self):
        region = self._cb_region.currentData()
        if region is None:
            return
        data = TABLE_2_5.get(region, {})
        tipo = self._cb_tipo_obra.currentData()
        bm = data.get("bm", 0)
        d50 = data.get("d50", 0)
        beta_ref = bm if tipo == OBRA_PM else bm - d50
        items = [f"{bm:.2f}", f"{d50:.2f}", "1.00", tipo]
        for col, val in enumerate(items):
            item = QTableWidgetItem(val)
            item.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
            self._tbl_beta.setItem(0, col, item)

    def _on_calc_tc(self):
        try:
            A = float(self._ed_area.text().replace(",", "."))
            KA = self._calc_KA(A)
            self._ed_KA.setText(f"{KA:.4f}")

            tipo_idx = self._cb_tipo_cuenca.currentIndex()
            if tipo_idx == 0:
                Lc = float(self._ed_lc.text().replace(",", "."))
                Jc = float(self._ed_jc.text().replace(",", "."))
                from .ic52ic_core import calc_tc_principal, calc_Kt
                tc = calc_tc_principal(Lc, Jc)
            else:
                Ldif = float(self._ed_ldif.text().replace(",", "."))
                ndif_map = list(NDIF.values())
                ndif = ndif_map[self._cb_ndif.currentIndex()]
                Jdif = float(self._ed_jdif.text().replace(",", "."))
                from .ic52ic_core import calc_tdif_min, apply_tc_table22, calc_Kt
                tdif = calc_tdif_min(Ldif, ndif, Jdif)
                tc = apply_tc_table22(tdif) / 60.0

            from .ic52ic_core import calc_Kt
            Kt = calc_Kt(tc)
            self._ed_tc.setText(f"{tc:.4f}")
            self._ed_Kt.setText(f"{Kt:.4f}")
        except (ValueError, ZeroDivisionError) as e:
            QMessageBox.warning(self, "Error en cálculo tc", str(e))

    def _calc_KA(self, A):
        import math
        if A < 1.0:
            return 1.0
        return 1.0 - math.log10(A) / 15.0

    def _on_calc_pd_cv(self):
        """Calculate Pd for all T from P_media and Cv, fill the Pd editors."""
        if not self._grp_cv.isChecked():
            return
        try:
            P_media = float(self._ed_P_media.text().replace(",", ".").strip())
            Cv      = float(self._ed_Cv.text().replace(",", ".").strip())
        except ValueError:
            QMessageBox.warning(self, "Datos incompletos", "Introduce P media y Cv válidos.")
            return

        from .ic52ic_p0table import calc_Pd_from_P_Cv, get_YT, CV_YT_TABLE

        cv_keys = sorted(CV_YT_TABLE.keys())
        if Cv < cv_keys[0] or Cv > cv_keys[-1]:
            QMessageBox.warning(
                self, "Cv fuera de rango",
                f"Cv = {Cv:.3f} fuera del rango de la tabla ({cv_keys[0]:.2f} – {cv_keys[-1]:.2f})."
            )
            return

        # T_key → editor mapping
        # MCO (2.5 years) interpolated between T=2 and T=5
        results = {}
        for T_key, ed in self._pd_editors.items():
            if T_key == "MCO":
                yt2 = get_YT(Cv, 2)
                yt5 = get_YT(Cv, 5)
                if yt2 is not None and yt5 is not None:
                    yt_mco = yt2 + (yt5 - yt2) * (2.5 - 2) / (5 - 2)
                    pd = round(P_media * yt_mco, 2)
                else:
                    pd = None
            else:
                pd = calc_Pd_from_P_Cv(P_media, Cv, T_key)
                if pd is not None:
                    pd = round(pd, 2)

            if pd is not None:
                ed.setText(str(pd))
                results[T_key] = pd
            else:
                ed.setText("")

        # Show summary
        lines = [f"P media = {P_media} mm  |  Cv = {Cv}", ""]
        for T_key, pd in results.items():
            label = "MCO" if T_key == "MCO" else f"T={T_key}"
            lines.append(f"  {label:8s}  Pd = {pd:.2f} mm")
        QMessageBox.information(
            self, "Pd calculados",
            "\n".join(lines) + "\n\nValores copiados a los campos de precipitación."
        )


    def _on_calc_p0_clc(self):
        layer = self._cb_capa_clc.currentLayer()
        if layer is None:
            QMessageBox.warning(self, "Capa CLC", "Selecciona una capa CLC válida.")
            return
        grupo_map = {0: "A", 1: "B", 2: "C", 3: "D"}
        grupo = grupo_map[self._cb_grupo_hid.currentIndex()]
        try:
            pend = float(self._ed_pendiente_umbral.text().replace(",", "."))
        except ValueError:
            pend = 3.0

        self._p0_worker = P0Worker(layer, grupo, pend, self)
        self._p0_worker.rows_ready.connect(self._on_p0_rows_ready)
        self._p0_worker.error.connect(lambda msg: QMessageBox.critical(self, "Error P0i", msg))
        self._p0_worker.start()

    def _on_p0_rows_ready(self, rows, total_area):
        """Populate the CLC breakdown table with one row per land use code."""
        from .ic52ic_p0table import CLC_NAMES
        self._tbl_clc.setRowCount(0)
        grupo_map = {0: "A", 1: "B", 2: "C", 3: "D"}
        grupo = grupo_map[self._cb_grupo_hid.currentIndex()]

        try:
            pend = float(self._ed_pendiente_umbral.text().replace(",", "."))
        except ValueError:
            pend = 3.0
        pend_label = f"{'≥' if pend >= 3.0 else '<'}3%"

        colors_ok  = QColor("#d5f5e3")
        colors_err = QColor("#fadbd8")

        for code_norm, code_raw, area, P0i, note in rows:
            row = self._tbl_clc.rowCount()
            self._tbl_clc.insertRow(row)
            bg = colors_ok if P0i is not None else colors_err

            # Col 0: Código CLC (normalizado)
            it0 = QTableWidgetItem(code_norm)
            it0.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
            it0.setBackground(bg)
            it0.setFlags(it0.flags() & ~(Qt.ItemFlag.ItemIsEditable if hasattr(Qt, "ItemFlag") else Qt.ItemIsEditable))
            self._tbl_clc.setItem(row, 0, it0)

            # Col 1: Nombre uso de suelo
            # code_norm may be padded (e.g. "24300") → resolve to canonical ("24310")
            from .ic52ic_p0table import get_clc_name
            name = get_clc_name(code_norm)
            it1 = QTableWidgetItem(name)
            it1.setBackground(bg)
            it1.setFlags(it1.flags() & ~(Qt.ItemFlag.ItemIsEditable if hasattr(Qt, "ItemFlag") else Qt.ItemIsEditable))
            self._tbl_clc.setItem(row, 1, it1)

            # Col 2: Grupo hidrológico
            it2 = QTableWidgetItem(grupo)
            it2.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
            it2.setBackground(bg)
            it2.setFlags(it2.flags() & ~(Qt.ItemFlag.ItemIsEditable if hasattr(Qt, "ItemFlag") else Qt.ItemIsEditable))
            self._tbl_clc.setItem(row, 2, it2)

            # Col 3: Pendiente asignada
            it3 = QTableWidgetItem(pend_label)
            it3.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
            it3.setBackground(bg)
            it3.setFlags(it3.flags() & ~(Qt.ItemFlag.ItemIsEditable if hasattr(Qt, "ItemFlag") else Qt.ItemIsEditable))
            self._tbl_clc.setItem(row, 3, it3)

            # Col 4: Área (km²)
            it4 = QTableWidgetItem(f"{area:.4f}")
            it4.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
            it4.setBackground(bg)
            it4.setFlags(it4.flags() & ~(Qt.ItemFlag.ItemIsEditable if hasattr(Qt, "ItemFlag") else Qt.ItemIsEditable))
            self._tbl_clc.setItem(row, 4, it4)

            # Col 5: P0i (mm) — EDITABLE
            p0i_val = f"{P0i:.1f}" if P0i is not None else "N/D"
            it5 = QTableWidgetItem(p0i_val)
            it5.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
            it5.setBackground(bg)
            if P0i is None:
                it5.setToolTip(f"Código {code_norm} no encontrado en Tabla 2.3 — introduce el valor manualmente")
            elif note:
                it5.setToolTip(f"⚠ {note}")
                it5.setForeground(QColor("#7d6608"))  # amber = approximated
            self._tbl_clc.setItem(row, 5, it5)

        self._ed_area_total_clc.setText(f"{total_area:.4f}")

    def _on_calcular(self):
        try:
            cuenca, precip, suelo = self._build_inputs()
        except ValueError as e:
            QMessageBox.warning(self, "Datos incompletos", str(e))
            return

        try:
            resultados = calcular_metodo_racional(
                cuenca, precip, suelo, periodos=PERIODOS_STANDARD
            )
        except Exception as e:
            QMessageBox.critical(self, "Error de cálculo", str(e))
            return

        self._resultados = resultados
        self._show_results(resultados)
        self._tabs.setCurrentIndex(3)

    def _build_inputs(self) -> tuple:
        """Parse dialog fields → CuencaInput, PrecipInput, SueloInput."""
        def _float(ed, name):
            txt = ed.text().replace(",", ".").strip()
            if not txt:
                raise ValueError(f"Campo '{name}' vacío.")
            return float(txt)

        # Cuenca
        A  = _float(self._ed_area, "Área A")
        Lc = _float(self._ed_lc,   "Longitud cauce Lc") if self._cb_tipo_cuenca.currentIndex() == 0 else 0.0
        Jc = _float(self._ed_jc,   "Pendiente Jc")      if self._cb_tipo_cuenca.currentIndex() == 0 else 0.0
        tipo = "principal" if self._cb_tipo_cuenca.currentIndex() == 0 else "secundaria"

        ldif = jdif = 0.0
        ndif_tipo = "bare_soil"
        if tipo == "secundaria":
            ldif = _float(self._ed_ldif, "Ldif")
            jdif = _float(self._ed_jdif, "Jdif")
            ndif_keys = list(NDIF.keys())
            ndif_tipo = ndif_keys[self._cb_ndif.currentIndex()]

        cuenca = CuencaInput(
            nombre=self._ed_nombre.text(),
            area_km2=A, longitud_cauce_km=Lc, pendiente_cauce=Jc,
            tipo_cuenca=tipo, ldif_m=ldif, ndif_tipo=ndif_tipo, jdif=jdif,
        )

        # Precipitación — read all 8 Pd values
        Pd_dict = {}
        for T_key, ed in self._pd_editors.items():
            txt = ed.text().replace(",", ".").strip()
            if not txt:
                raise ValueError(
                    f"Campo Pd vacío para T={T_key}. "
                    f"Introduce todos los valores de precipitación diaria."
                )
            Pd_dict[T_key] = float(txt)

        I1Id  = _float(self._ed_I1Id,  "I1/Id")
        region = self._cb_region.currentData()
        tipo_obra = self._cb_tipo_obra.currentData()

        usar_idf = self._grp_idf.isChecked()
        IIDF_tc = IIDF_24 = None
        kb = 1.13
        if usar_idf:
            IIDF_tc = _float(self._ed_IIDF_tc, "I_IDF(T,tc)")
            IIDF_24 = _float(self._ed_IIDF_24, "I_IDF(T,24)")
            try:
                kb = float(self._ed_kb.text().replace(",", "."))
            except ValueError:
                kb = 1.13

        precip = PrecipInput(
            Pd=Pd_dict,
            I1_Id=I1Id, region_beta=region, tipo_obra=tipo_obra,
            usar_IDF=usar_idf, IIDF_tc=IIDF_tc, IIDF_24=IIDF_24, kb=kb,
        )

        # Suelo
        grupo_map = {0: "A", 1: "B", 2: "C", 3: "D"}
        grupo = grupo_map[self._cb_grupo_hid.currentIndex()]

        # Collect CLC breakdown rows from table
        clc_rows = []
        if self._tbl_clc.rowCount() > 0 and not self._grp_p0_manual.isChecked():
            for row in range(self._tbl_clc.rowCount()):
                code = self._tbl_clc.item(row, 0).text()
                area_txt = self._tbl_clc.item(row, 4).text().replace(",", ".")
                p0i_txt  = self._tbl_clc.item(row, 5).text().replace(",", ".")
                if p0i_txt in ("N/D", ""):
                    raise ValueError(
                        f"El código CLC {code} tiene P0i sin determinar (N/D). "
                        f"Introduce el valor manualmente en la celda correspondiente."
                    )
                clc_rows.append({
                    "code": code,
                    "area_km2": float(area_txt),
                    "P0i": float(p0i_txt),
                })
            suelo = SueloInput(P0i=None, grupo_hid=grupo, clc_rows=clc_rows)
        else:
            # Manual override
            p0i_txt = self._ed_P0i_manual.text().replace(",", ".").strip()
            if not p0i_txt:
                raise ValueError(
                    "No hay tabla CLC cargada ni P0i manual introducido. "
                    "Carga la capa CLC o activa la opción manual."
                )
            suelo = SueloInput(P0i=float(p0i_txt), grupo_hid=grupo, clc_rows=[])

        return cuenca, precip, suelo

    # ------------------------------------------------------------------
    # Show results
    # ------------------------------------------------------------------

    def _show_results(self, resultados):
        if not resultados:
            return
        r0 = resultados[0]

        # Intermediate values (common)
        self._res_eds["KA"].setText(f"{r0.KA:.4f}")
        self._res_eds["tc (h)"].setText(f"{r0.tc_h:.4f}")
        self._res_eds["Kt"].setText(f"{r0.Kt:.4f}")
        self._res_eds["I₁/Id"].setText(f"{r0.Fa:.4f}")
        self._res_eds["P0i (mm)"].setText(f"{r0.P0i:.2f}")

        # Results table
        self._tbl_res.setRowCount(len(resultados))

        T_to_Pd = {T_key: ed.text() for T_key, ed in self._pd_editors.items()}

        # Row color gradient: blue → green → orange → red
        row_colors = {
            2:     "#d6eaf8",
            "MCO": "#c8e6f9",
            5:     "#d5f5e3",
            10:    "#d5f5e3",
            25:    "#fef9e7",
            50:    "#fdebd0",
            100:   "#fadbd8",
            500:   "#f2d7d5",
        }

        for row, res in enumerate(resultados):
            T_key = res.T
            T_label = "MCO" if T_key == "MCO" else str(T_key)
            Pd_val = T_to_Pd.get(T_key, "")
            bg = QColor(row_colors.get(T_key, "#ffffff"))

            vals = [
                T_label,
                Pd_val,
                f"{res.Id_mmh:.3f}",
                f"{res.Fint:.3f}",
                f"{res.I_mmh:.3f}",
                f"{res.beta:.4f}",
                f"{res.P0:.2f}",
                f"{res.C:.4f}",
                f"{res.suma_CiAi:.4f}" if res.heterogeneo else f"{res.C * res.KA:.4f}",
                f"{res.QT_m3s:.3f}",
            ]
            for col, val in enumerate(vals):
                item = QTableWidgetItem(val)
                item.setTextAlignment((Qt.AlignmentFlag.AlignCenter if hasattr(Qt, "AlignmentFlag") else Qt.AlignCenter))
                item.setBackground(bg)
                if col == 9:  # QT column
                    f = item.font()
                    f.setBold(True)
                    item.setFont(f)
                self._tbl_res.setItem(row, col, item)

        # QT highlight boxes
        T_map = {r.T: r.QT_m3s for r in resultados}
        for T_key, ed in self._q_boxes.items():
            val = T_map.get(T_key)
            ed.setText(f"{val:.3f}" if val is not None else "—")

        # Warnings
        all_warns = []
        for res in resultados:
            for w in res.warnings:
                all_warns.append(f"T={res.T}: {w}")
        self._txt_warnings.setPlainText("\n".join(all_warns) if all_warns else "Sin avisos.")

    def _on_export_excel(self):
        """Export results table and intermediate variables to .xlsx."""
        if not self._resultados:
            QMessageBox.warning(self, "Sin resultados", "Primero calcula los caudales.")
            return

        # Ask for save path
        path, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel", "IC52IC_Resultados.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(
                self, "openpyxl no disponible",
                "Instala openpyxl: en la consola Python de QGIS ejecuta\n"
                "import subprocess; subprocess.run(['pip', 'install', 'openpyxl'])"
            )
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "IC52IC Resultados"

            # ── Styles ────────────────────────────────────────────────────────
            hdr_font   = Font(bold=True, color="FFFFFF", size=10)
            hdr_fill_b = PatternFill("solid", fgColor="1A5276")  # blue header
            hdr_fill_g = PatternFill("solid", fgColor="1E8449")  # green sub-header
            center     = Alignment(horizontal="center", vertical="center")
            thin       = Side(style="thin", color="CCCCCC")
            border     = Border(left=thin, right=thin, top=thin, bottom=thin)

            T_COLORS = {
                2:   "D6EAF8", "MCO": "D6EAF8",
                5:   "D5F5E3", 10:  "D5F5E3",
                25:  "FEF9E7", 50:  "FEF9E7",
                100: "FDEBD0", 500: "FADBD8",
            }

            def hdr(ws, row, col, text, fill=None, font=None, span=1):
                cell = ws.cell(row=row, column=col, value=text)
                cell.font  = font or Font(bold=True, size=10)
                cell.fill  = fill or PatternFill("solid", fgColor="F2F3F4")
                cell.alignment = center
                cell.border = border
                if span > 1:
                    ws.merge_cells(
                        start_row=row, start_column=col,
                        end_row=row, end_column=col+span-1
                    )
                return cell

            def val(ws, row, col, v, fmt=None, fill=None):
                cell = ws.cell(row=row, column=col, value=v)
                cell.alignment = center
                cell.border = border
                if fmt:
                    cell.number_format = fmt
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
                return cell

            r0 = self._resultados[0]
            nombre = self._ed_nombre.text() or "Sin nombre"

            # ── Title ─────────────────────────────────────────────────────────
            ws.merge_cells("A1:J1")
            t = ws["A1"]
            t.value = f"IC 5.2-IC — Método Racional | {nombre}"
            t.font = Font(bold=True, size=12, color="FFFFFF")
            t.fill = PatternFill("solid", fgColor="1A5276")
            t.alignment = center

            # ── Common intermediate variables ─────────────────────────────────
            ws.merge_cells("A2:J2")
            ws["A2"].value = (
                f"KA = {r0.KA:.4f}   |   tc = {r0.tc_h:.4f} h   |   "
                f"Kt = {r0.Kt:.4f}   |   I₁/Id = {r0.Fa:.3f}   |   "
                f"P0i = {r0.P0i:.2f} mm   |   "
                f"{'Cuenca heterogénea §2.2.4' if r0.heterogeneo else 'Cuenca homogénea'}"
            )
            ws["A2"].font = Font(italic=True, size=9)
            ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

            # ── Results table header ──────────────────────────────────────────
            ROW_H = 3
            cols = [
                "T (años)", "Pd (mm)", "Id (mm/h)", "Fint", "I (mm/h)",
                "β", "P0 (mm)", "C", "Σ(Ci·Ai)", "QT (m³/s)"
            ]
            for c, name in enumerate(cols, 1):
                hdr(ws, ROW_H, c, name, fill=hdr_fill_b, font=hdr_font)

            # ── Data rows ────────────────────────────────────────────────────
            T_to_Pd = {}
            try:
                _, precip, _ = self._build_inputs()
                T_to_Pd = precip.Pd
            except Exception:
                pass

            for i, res in enumerate(self._resultados):
                row = ROW_H + 1 + i
                T_key = res.T
                fill = T_COLORS.get(T_key, "FFFFFF")
                t_label = "MCO" if T_key == "MCO" else str(T_key)
                pd_val = T_to_Pd.get(T_key, "")
                sigma = res.suma_CiAi if res.heterogeneo else res.C * r0.KA

                val(ws, row, 1,  t_label,      fill=fill)
                val(ws, row, 2,  pd_val,        "0.00", fill=fill)
                val(ws, row, 3,  res.Id_mmh,    "0.000", fill=fill)
                val(ws, row, 4,  res.Fint,      "0.000", fill=fill)
                val(ws, row, 5,  res.I_mmh,     "0.000", fill=fill)
                val(ws, row, 6,  res.beta,      "0.0000", fill=fill)
                val(ws, row, 7,  res.P0,        "0.00", fill=fill)
                val(ws, row, 8,  res.C,         "0.0000", fill=fill)
                val(ws, row, 9,  sigma,         "0.0000", fill=fill)
                qt_cell = val(ws, row, 10, res.QT_m3s, "0.000", fill=fill)
                qt_cell.font = Font(bold=True, size=10)

            # ── QT summary ────────────────────────────────────────────────────
            ROW_S = ROW_H + len(self._resultados) + 2
            ws.merge_cells(
                start_row=ROW_S, start_column=1, end_row=ROW_S, end_column=10
            )
            s = ws.cell(row=ROW_S, column=1, value="CAUDALES DE DISEÑO QT (m³/s)")
            s.font = Font(bold=True, color="FFFFFF", size=10)
            s.fill = PatternFill("solid", fgColor="1E8449")
            s.alignment = center

            ROW_S2 = ROW_S + 1
            for c, res in enumerate(self._resultados, 1):
                t_label = "MCO" if res.T == "MCO" else f"T={res.T}"
                fill = T_COLORS.get(res.T, "FFFFFF")
                hdr(ws, ROW_S2,   c, t_label, fill=PatternFill("solid", fgColor=fill))
                val(ws, ROW_S2+1, c, res.QT_m3s, "0.000", fill=fill)
                ws.cell(row=ROW_S2+1, column=c).font = Font(bold=True, size=11)

            # ── Column widths ─────────────────────────────────────────────────
            for c in range(1, 11):
                ws.column_dimensions[get_column_letter(c)].width = 13
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 18

            wb.save(path)
            QMessageBox.information(
                self, "Excel exportado",
                f"Archivo guardado en:\n{path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Error al exportar", str(e))

    def _on_cerrar(self):
        self._cleanup_map_layers()
        self.reject()

    def closeEvent(self, event):
        self._cleanup_map_layers()
        super().closeEvent(event)

    def _on_limpiar(self):
        self._tbl_clc.setRowCount(0)
        self._ed_area_total_clc.clear()
        for ed in [self._ed_nombre, self._ed_area, self._ed_lc, self._ed_jc,
                   self._ed_KA, self._ed_tc, self._ed_Kt,
                   self._ed_I1Id, self._ed_P0i_manual,
                   self._ed_P_media, self._ed_Cv,
                   self._ed_area_total_clc]:
            ed.clear()
        for ed in self._pd_editors.values():
            ed.clear()
        self._tbl_res.setRowCount(0)
        for ed in self._q_boxes.values():
            ed.clear()
        for ed in self._res_eds.values():
            ed.clear()
        self._txt_warnings.clear()
